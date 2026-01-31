"""
ChemExtract Web Application - Firebase-ready version
Không sử dụng win32com, có thể deploy trên Firebase/Linux
"""
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from werkzeug.utils import secure_filename
import os
import re
import json
from pathlib import Path
from filter import SDSCASReader
from datetime import datetime
import logging
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from copy import copy
import time
from concurrent.futures import ThreadPoolExecutor, as_completed

# Cấu hình logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__, static_folder='.')
CORS(app)

# Cấu hình
UPLOAD_FOLDER = 'uploads'
RESULTS_FOLDER = 'results'
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg'}
MAX_FILES = 50
EXCEL_FILE = 'CHECK DANH MUC HOA CHAT.xlsx'
CAS_DATABASE_FILE = 'cas_database.json'

# Tạo thư mục nếu chưa tồn tại
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULTS_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max

# Load CAS database từ JSON
CAS_DATABASE = {}
CATEGORY_MAPPING = {
    'HC CÓ ĐIỀU KIỆN': 'HC CÓ ĐIỀU KIỆN',
    'HC KIỂM SOÁT ĐẶC BIỆT NHÓM 1': 'HC KIỂM SOÁT ĐẶC BIỆT 1',
    'HC KIỂM SOÁT ĐẶC BIỆT NHÓM 2': 'HC KIỂM SOÁT ĐẶC BIỆT 2',
    'HC CÓ KHPN': 'HC CÓ KHPN',
    'TIỀN CHẤT THUỐC NỔ': 'TIỀN CHẤT THUỐC NỔ',
    'SUY GIẢM TẦNG OZONE': 'SUY GIẢM TẦNG OZONE',
    'HC CẤM': 'HC CẤM',
    'HC BẢNG 1': 'HC BẢNG 1'
}


def load_cas_database():
    """Load CAS database từ JSON file"""
    global CAS_DATABASE
    try:
        if os.path.exists(CAS_DATABASE_FILE):
            with open(CAS_DATABASE_FILE, 'r', encoding='utf-8') as f:
                CAS_DATABASE = json.load(f)
            logger.info(f"Loaded CAS database: {sum(len(v) for v in CAS_DATABASE.values())} total CAS numbers")
        else:
            logger.warning(f"CAS database file not found: {CAS_DATABASE_FILE}")
    except Exception as e:
        logger.error(f"Error loading CAS database: {e}")


def check_cas_in_database(cas_number):
    """
    Kiểm tra CAS number có trong các danh mục hay không.
    Trả về dict với kết quả cho mỗi danh mục.
    
    Logic tương đương với công thức Excel:
    =IF(COUNTIF('HC CÓ ĐIỀU KIỆN'!$A$5:$E$799, B2) > 0, "X", "")
    """
    results = {
        'HC CÓ ĐIỀU KIỆN': '',
        'HC KIỂM SOÁT ĐẶC BIỆT 1': '',
        'HC KIỂM SOÁT ĐẶC BIỆT 2': '',
        'HC CÓ KHPN': '',
        'TIỀN CHẤT THUỐC NỔ': '',
        'SUY GIẢM TẦNG OZONE': '',
        'HC CẤM': '',
        'HC BẢNG 1': ''
    }
    
    for db_key, result_key in CATEGORY_MAPPING.items():
        if db_key in CAS_DATABASE:
            if cas_number in CAS_DATABASE[db_key]:
                results[result_key] = 'X'
    
    return results


def allowed_file(filename):
    """Kiểm tra file extension hợp lệ"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/')
def index():
    """Serve trang chủ"""
    return send_from_directory('.', 'index.html')


@app.route('/index.html')
def index_html():
    """Serve trang chủ (explicit path)"""
    return send_from_directory('.', 'index.html')


@app.route('/sign_in.html')
def sign_in():
    """Serve trang đăng nhập"""
    return send_from_directory('.', 'sign_in.html')


@app.route('/<path:filename>')
def serve_static(filename):
    """Serve các file static khác"""
    return send_from_directory('.', filename)


@app.route('/upload', methods=['POST'])
def upload_files():
    """Xử lý upload files"""
    try:
        if 'files' not in request.files:
            return jsonify({'error': 'Không tìm thấy files'}), 400

        files = request.files.getlist('files')

        if len(files) > MAX_FILES:
            return jsonify({'error': f'Tối đa {MAX_FILES} files'}), 400

        uploaded_files = []
        for file in files:
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                # Thêm timestamp để tránh trùng lặp
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"{timestamp}_{filename}"
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                uploaded_files.append({
                    'filename': filename,
                    'original_name': file.filename,
                    'path': filepath
                })
            else:
                return jsonify({'error': f'File không hợp lệ: {file.filename}'}), 400

        return jsonify({
            'message': f'Upload thành công {len(uploaded_files)} files',
            'files': uploaded_files
        }), 200

    except Exception as e:
        logger.error(f"Upload error: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/process', methods=['POST'])
def process_files():
    """Xử lý files và trích xuất CAS numbers - không cần win32com"""
    try:
        total_start = time.time()

        data = request.json
        files = data.get('files', [])

        if not files:
            return jsonify({'error': 'Không có files để xử lý'}), 400

        # Danh sách các dict {product_name, cas}
        all_cas_entries = []

        # ===== PHASE 1: Trích xuất CAS từ PDF (song song) =====
        phase1_start = time.time()

        def process_single_pdf(file_info):
            """Xử lý một file PDF - dùng cho thread pool"""
            reader = SDSCASReader()  # Mỗi thread cần instance riêng
            filepath = file_info.get('path')
            original_name = file_info.get('original_name')

            if not os.path.exists(filepath):
                return None

            try:
                if filepath.lower().endswith('.pdf'):
                    # Trích xuất text từ PDF
                    text = reader.extract_text_from_pdf(filepath)

                    # Tìm CAS numbers theo thứ tự xuất hiện
                    cas_data = reader.find_cas_numbers(text)
                    valid_cas = cas_data.get('valid', [])

                    # Tên sản phẩm lấy từ tên file
                    product_name = os.path.splitext(original_name)[0]

                    return {
                        'product_name': product_name,
                        'cas_list': valid_cas,
                        'original_name': original_name
                    }
            except Exception as e:
                logger.error(f"Error processing {original_name}: {str(e)}")
            return None

        # Xử lý song song với ThreadPoolExecutor
        results = []
        max_workers = min(len(files), 4)  # Tối đa 4 threads

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # Submit tất cả tasks
            future_to_file = {executor.submit(
                process_single_pdf, f): f for f in files}

            # Thu thập kết quả theo thứ tự submit
            for future in future_to_file:
                result = future.result()
                if result:
                    results.append(result)

        # Sắp xếp lại theo thứ tự file gốc
        file_order = {f.get('original_name'): i for i, f in enumerate(files)}
        results.sort(key=lambda x: file_order.get(x['original_name'], 999))

        # Chuyển kết quả thành all_cas_entries
        for result in results:
            product_name = result['product_name']
            valid_cas = result['cas_list']
            original_name = result['original_name']

            logger.info(
                f"Found {len(valid_cas)} CAS numbers in {original_name}")

            for idx, cas in enumerate(valid_cas):
                entry = {
                    'product_name': product_name if idx == 0 else '',
                    'cas': cas
                }
                all_cas_entries.append(entry)
                if idx == 0:
                    logger.info(f"First CAS entry for {product_name}: {entry}")

        phase1_time = time.time() - phase1_start
        logger.info(
            f"⏱️ PHASE 1 - Trích xuất PDF (parallel): {phase1_time:.2f}s")

        if not all_cas_entries:
            return jsonify({'error': 'Không tìm thấy CAS numbers nào'}), 400

        # ===== PHASE 2: Kiểm tra CAS trong database (thay thế win32com) =====
        phase2_start = time.time()
        
        num_cas = len(all_cas_entries)
        table_data = []
        
        for entry in all_cas_entries:
            cas_number = entry['cas']
            
            # Kiểm tra CAS trong database - logic tương đương công thức Excel
            check_results = check_cas_in_database(cas_number)
            
            row_data = {
                'TÊN SẢN PHẨM': entry['product_name'],
                'CAS NUMBER': cas_number,
                'HC CÓ ĐIỀU KIỆN': check_results['HC CÓ ĐIỀU KIỆN'],
                'HC KIỂM SOÁT ĐẶC BIỆT 1': check_results['HC KIỂM SOÁT ĐẶC BIỆT 1'],
                'HC KIỂM SOÁT ĐẶC BIỆT 2': check_results['HC KIỂM SOÁT ĐẶC BIỆT 2'],
                'HC CÓ KHPN': check_results['HC CÓ KHPN'],
                'TIỀN CHẤT THUỐC NỔ': check_results['TIỀN CHẤT THUỐC NỔ'],
                'SUY GIẢM TẦNG OZONE': check_results['SUY GIẢM TẦNG OZONE'],
                'HC CẤM': check_results['HC CẤM'],
                'HC BẢNG 1': check_results['HC BẢNG 1']
            }
            table_data.append(row_data)
        
        phase2_time = time.time() - phase2_start
        logger.info(f"⏱️ PHASE 2 - Kiểm tra CAS database: {phase2_time:.2f}s")

        # ===== PHASE 3: Tạo file Excel kết quả =====
        phase3_start = time.time()
        
        # Tạo workbook mới với kết quả
        wb = Workbook()
        ws = wb.active
        ws.title = "CHECK CAS"
        
        # Headers
        headers = ['TÊN SẢN PHẨM', 'CAS NUMBER', 'HC CÓ ĐIỀU KIỆN',
                   'HC KIỂM SOÁT ĐẶC BIỆT 1', 'HC KIỂM SOÁT ĐẶC BIỆT 2',
                   'HC CÓ KHPN', 'TIỀN CHẤT THUỐC NỔ', 'SUY GIẢM TẦNG OZONE',
                   'HC CẤM', 'HC BẢNG 1']
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Ghi dữ liệu
        for i, row_data in enumerate(table_data):
            row = i + 2
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=row_data.get(header, ''))
        
        # Tạo file backup với timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_filename = f'CAS_Extract_{timestamp}.xlsx'
        excel_path = os.path.join(RESULTS_FOLDER, excel_filename)
        wb.save(excel_path)
        
        phase3_time = time.time() - phase3_start
        logger.info(f"⏱️ PHASE 3 - Tạo file Excel: {phase3_time:.2f}s")

        # Log để debug
        logger.info(f"Total entries in table_data: {len(table_data)}")
        for i, row in enumerate(table_data):
            if row.get('TÊN SẢN PHẨM'):
                logger.info(
                    f"Row {i}: Product name = {row.get('TÊN SẢN PHẨM')}")

        # Tạo new_rows để highlight
        new_rows = []
        for entry in all_cas_entries:
            new_rows.append({
                'TÊN SẢN PHẨM': entry['product_name'],
                'CAS NUMBER': entry['cas']
            })

        total_time = time.time() - total_start
        logger.info(f"")
        logger.info(f"{'='*50}")
        logger.info(f"📊 TỔNG KẾT THỜI GIAN XỬ LÝ")
        logger.info(f"{'='*50}")
        logger.info(f"   PHASE 1 - Trích xuất PDF:        {phase1_time:.2f}s")
        logger.info(f"   PHASE 2 - Kiểm tra CAS database: {phase2_time:.2f}s")
        logger.info(f"   PHASE 3 - Tạo file Excel:        {phase3_time:.2f}s")
        logger.info(f"{'='*50}")
        logger.info(f"   🕐 TỔNG THỜI GIAN:               {total_time:.2f}s")
        logger.info(f"{'='*50}")
        logger.info(f"")

        return jsonify({
            'message': f'Trích xuất thành công {num_cas} CAS numbers',
            'total_cas': num_cas,
            'total_files': len(files),
            'excel_file': excel_filename,
            'table_data': table_data,
            'new_rows': new_rows
        }), 200

    except Exception as e:
        logger.error(f"Process error: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/download/<filename>')
def download_file(filename):
    """Download file Excel kết quả"""
    try:
        return send_file(
            os.path.join(RESULTS_FOLDER, filename),
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logger.error(f"Download error: {str(e)}")
        return jsonify({'error': 'File không tồn tại'}), 404


@app.route('/cleanup', methods=['POST'])
def cleanup_files():
    """Dọn dẹp files tạm"""
    try:
        # Xóa uploaded files
        for filename in os.listdir(UPLOAD_FOLDER):
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            if os.path.isfile(filepath):
                os.remove(filepath)

        return jsonify({'message': 'Dọn dẹp thành công'}), 200
    except Exception as e:
        logger.error(f"Cleanup error: {str(e)}")
        return jsonify({'error': str(e)}), 500


# Load CAS database khi khởi động
load_cas_database()

if __name__ == '__main__':
    print("🚀 Starting ChemExtract Server (Firebase-ready version)...")
    print("📍 Server running at: http://localhost:5000")
    print("📄 Open your browser and go to: http://localhost:5000")
    print("✅ No win32com dependency - can deploy on Linux/Firebase!")
    app.run(debug=True, host='0.0.0.0', port=5000)
