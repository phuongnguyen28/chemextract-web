import re
import pandas as pd
import pdfplumber
from pathlib import Path
import os
import config

# Thử import PyMuPDF (fitz) - nhanh hơn pdfplumber
try:
    import fitz  # PyMuPDF
    HAS_PYMUPDF = True
except ImportError:
    HAS_PYMUPDF = False


class SDSCASReader:
    def __init__(self):
        # Regex pattern để nhận diện CAS Number
        self.cas_pattern = r'\b[1-9]\d{1,6}-\d{2}-\d\b'
        self.cas_numbers_found = []

        # Các từ khóa để phát hiện Điều 8 (hazardous materials)
        self.dieu_8_keywords = [
            "Hazardous material, 1",
            "Hazardous material, 2",
            "Hazardous material, 3",
            "Acute toxicity, Category 1",
            "Acute toxicity, Category 2",
            "Acute toxicity, Category 3",
            "Acute Tox. 1",
            "Acute Tox. 2",
            "Acute Tox. 3",
            "Serious eye damage, Category 1",
            "Serious eye damage, Category 2",
            "Serious eye damage, Category 2A",
            "Eye Dam. 1",
            "Eye Dam. 2",
            "Eye Dam. 2A",
            "Eye irritation, Category 1",
            "Eye irritation, Category 2",
            "Eye irritation, Category 2A",
            "Eye Irrit. 1",
            "Eye Irrit. 2",
            "Eye Irrit. 2A",
            "Skin irritation, Category 1",
            "Skin irritation, Category 2",
            "Skin Irrit. 1",
            "Skin Irrit. 2",
            "Carcinogenicity, Category 2",
            "Germ cell mutagenicity, Category 2",
            "Reproductive toxicity, Category 2",
            "Environmental hazard, Category 1"
        ]

        # Các từ khóa để phát hiện Điều 14 (severe hazards)
        self.dieu_14_keywords = [
            "Acute toxicity, Category 1",
            "Acute Tox. 1",
            "Carcinogenicity, Category 1A",
            "Carcinogenicity, Category 1B",
            "Reproductive toxicity, Category 1A",
            "Reproductive toxicity, Category 1B",
            "Germ cell mutagenicity, Category 1A",
            "Germ cell mutagenicity, Category 1B"
        ]

    def extract_text_from_pdf(self, pdf_path):
        """Trích xuất văn bản từ file PDF - ưu tiên PyMuPDF (nhanh hơn)"""
        text = ""

        # Ưu tiên dùng PyMuPDF (fitz) - nhanh hơn 5-10x so với pdfplumber
        if HAS_PYMUPDF:
            try:
                doc = fitz.open(pdf_path)
                for page in doc:
                    # Dùng sort=True để giữ thứ tự text từ trên xuống dưới
                    # flags=fitz.TEXT_PRESERVE_WHITESPACE giữ nguyên khoảng trắng
                    text += page.get_text("text", sort=True) + "\n"
                doc.close()
                return text
            except Exception as e:
                # Fallback to pdfplumber nếu PyMuPDF lỗi
                pass

        # Fallback: dùng pdfplumber
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
            return text
        except Exception as e:
            raise Exception(f"Lỗi khi đọc file PDF: {str(e)}")

    def find_cas_numbers(self, text):
        """Tìm tất cả CAS numbers trong văn bản và phân loại - GIỮ NGUYÊN THỨ TỰ"""
        # Tìm tất cả matches - giữ nguyên thứ tự xuất hiện
        matches = re.findall(self.cas_pattern, text)

        valid_cas_numbers = []
        invalid_cas_numbers = []
        all_cas_ordered = []  # Danh sách tất cả CAS theo thứ tự xuất hiện
        seen_valid = set()
        seen_invalid = set()
        seen_all = set()

        # Giữ nguyên thứ tự xuất hiện, chỉ loại bỏ duplicate
        for match in matches:
            is_valid = self.validate_cas_number(match)

            # Thêm vào danh sách tổng theo thứ tự xuất hiện
            if match not in seen_all:
                all_cas_ordered.append({'cas': match, 'valid': is_valid})
                seen_all.add(match)

            if is_valid:
                if match not in seen_valid:
                    valid_cas_numbers.append(match)
                    seen_valid.add(match)
            else:
                if match not in seen_invalid:
                    invalid_cas_numbers.append(match)
                    seen_invalid.add(match)

        return {
            'valid': valid_cas_numbers,  # Giữ nguyên thứ tự, đã loại bỏ duplicate
            'invalid': invalid_cas_numbers,  # Giữ nguyên thứ tự, đã loại bỏ duplicate
            'all_ordered': all_cas_ordered  # Tất cả CAS theo thứ tự xuất hiện trong SDS
        }

    def validate_cas_number(self, cas_number):
        """Xác thực CAS number bằng checksum"""
        try:
            # Loại bỏ dấu gạch ngang
            digits = cas_number.replace('-', '')

            # Tính checksum
            checksum = int(digits[-1])
            digits = digits[:-1]

            total = 0
            for i, digit in enumerate(reversed(digits)):
                total += int(digit) * (i + 1)

            return total % 10 == checksum
        except:
            return False

    def detect_dieu_8_hazards(self, text):
        """Phát hiện các chất nguy hiểm theo Điều 8"""
        text_lower = text.lower()
        for keyword in self.dieu_8_keywords:
            if keyword.lower() in text_lower:
                return True
        return False

    def detect_dieu_14_hazards(self, text):
        """Phát hiện các chất nguy hiểm theo Điều 14"""
        text_lower = text.lower()
        for keyword in self.dieu_14_keywords:
            if keyword.lower() in text_lower:
                return True
        return False

    # Aliases for compatibility with app.py
    def detect_dieu_8(self, text):
        """Alias for detect_dieu_8_hazards"""
        return self.detect_dieu_8_hazards(text)

    def detect_dieu_14(self, text):
        """Alias for detect_dieu_14_hazards"""
        return self.detect_dieu_14_hazards(text)

    def process_sds_file(self, pdf_path):
        """Xử lý một file SDS"""
        try:
            print(f"Đang xử lý: {os.path.basename(pdf_path)}")
            text = self.extract_text_from_pdf(pdf_path)
            cas_results = self.find_cas_numbers(text)

            return {
                'file_name': os.path.basename(pdf_path),
                'valid_cas_numbers': cas_results['valid'],
                'invalid_cas_numbers': cas_results['invalid'],
                # Thêm danh sách theo thứ tự
                'all_cas_ordered': cas_results['all_ordered'],
                'count_valid': len(cas_results['valid']),
                'count_invalid': len(cas_results['invalid'])
            }
        except Exception as e:
            return {
                'file_name': os.path.basename(pdf_path),
                'valid_cas_numbers': [],
                'invalid_cas_numbers': [],
                'all_cas_ordered': [],  # Thêm danh sách rỗng khi lỗi
                'count_valid': 0,
                'count_invalid': 0,
                'error': str(e)
            }

    def process_multiple_files(self, pdf_paths):
        """Xử lý nhiều file SDS"""
        results = []

        for pdf_path in pdf_paths:
            result = self.process_sds_file(pdf_path)
            results.append(result)

            # Thêm tất cả CAS numbers hợp lệ vào danh sách tổng
            self.cas_numbers_found.extend(result['valid_cas_numbers'])

        # Loại bỏ duplicates và sắp xếp
        self.cas_numbers_found = sorted(list(set(self.cas_numbers_found)))

        return results

    def export_to_excel(self, results):
        """Xuất kết quả ra file Excel - giữ nguyên cấu trúc và công thức"""
        try:
            from openpyxl import load_workbook, Workbook
            from openpyxl.utils import get_column_letter

            # Tạo danh sách dữ liệu mới từ kết quả SDS
            new_data = []

            for result in results:
                file_name = result['file_name']
                # Lấy tên file không có đuôi .pdf
                product_name = os.path.splitext(file_name)[0]
                all_cas_ordered = result.get('all_cas_ordered', [])

                # Lọc chỉ lấy CAS hợp lệ
                valid_cas_only = [
                    cas_info for cas_info in all_cas_ordered if cas_info.get('valid', False)]

                # Nếu có CAS hợp lệ
                if len(valid_cas_only) > 0:
                    # Thêm từng CAS hợp lệ theo đúng thứ tự xuất hiện trong SDS
                    for i, cas_info in enumerate(valid_cas_only):
                        cas = cas_info['cas']
                        new_data.append({
                            'TÊN SẢN PHẨM': product_name if i == 0 else "",
                            'CAS': cas
                        })
                else:
                    # Nếu không có CAS hợp lệ
                    new_data.append({
                        'TÊN SẢN PHẨM': product_name,
                        'CAS': ""
                    })

            # Mở file Excel hiện có bằng openpyxl để giữ nguyên công thức
            if os.path.exists(config.EXCEL_SOURCE):
                wb = load_workbook(config.EXCEL_SOURCE)
                ws = wb.active

                # Tìm header columns
                header_row = 1
                col_mapping = {}
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=header_row, column=col).value
                    if cell_value:
                        col_mapping[cell_value] = col

                # Tìm cột CAS
                cas_col = None
                if 'CAS' in col_mapping:
                    cas_col = col_mapping['CAS']
                elif 'CAS NUMBER' in col_mapping:
                    cas_col = col_mapping['CAS NUMBER']

                # Tìm cột TÊN SẢN PHẨM
                ten_sp_col = col_mapping.get('TÊN SẢN PHẨM')

                # XÓA TOÀN BỘ DỮ LIỆU CŨ Ở 2 CỘT TÊN SẢN PHẨM VÀ CAS (từ dòng 2 trở đi)
                # Tìm dòng cuối cùng thực sự có dữ liệu
                max_data_row = ws.max_row
                # Tăng thêm để đảm bảo xóa hết (có thể có dòng ẩn hoặc trống)
                for extra_row in range(max_data_row + 1, max_data_row + 1000):
                    has_data = False
                    for col in range(1, ws.max_column + 1):
                        if ws.cell(row=extra_row, column=col).value:
                            has_data = True
                            max_data_row = extra_row
                            break
                    if not has_data:
                        break

                # Xóa từ dòng 2 đến dòng cuối cùng - Dùng chuỗi rỗng thay vì None
                if max_data_row > header_row:
                    for row in range(header_row + 1, max_data_row + 1):
                        # Xóa TÊN SẢN PHẨM
                        if ten_sp_col:
                            ws.cell(row=row, column=ten_sp_col).value = ""
                        # Xóa CAS
                        if cas_col:
                            ws.cell(row=row, column=cas_col).value = ""

                # Bắt đầu ghi từ dòng 2
                last_row = header_row

                # Thêm dữ liệu mới
                for row_data in new_data:
                    last_row += 1

                    # Ghi TÊN SẢN PHẨM
                    if ten_sp_col:
                        ws.cell(row=last_row, column=ten_sp_col,
                                value=row_data.get('TÊN SẢN PHẨM', ''))

                    # Ghi CAS
                    if cas_col:
                        ws.cell(row=last_row, column=cas_col,
                                value=row_data.get('CAS', ''))

                # Lưu file
                wb.save(config.EXCEL_SOURCE)
                wb.close()

            else:
                # Nếu file không tồn tại, tạo mới bằng openpyxl
                wb = Workbook()
                ws = wb.active

                # Tạo header
                headers = ['TÊN SẢN PHẨM', 'CAS']
                for col_idx, header in enumerate(headers, start=1):
                    ws.cell(row=1, column=col_idx, value=header)

                # Thêm dữ liệu
                for row_idx, row_data in enumerate(new_data, start=2):
                    ws.cell(row=row_idx, column=1,
                            value=row_data.get('TÊN SẢN PHẨM', ''))
                    ws.cell(row=row_idx, column=2,
                            value=row_data.get('CAS', ''))

                # Lưu file
                wb.save(config.EXCEL_SOURCE)
                wb.close()

            return True
        except Exception as e:
            raise Exception(f"Lỗi khi xuất file Excel: {str(e)}")


class SDSApp:
    def __init__(self, root):
        self.root = root
        self.root.title("SDS CAS Number Extractor")
        self.root.geometry("450x300")

        self.reader = SDSCASReader()
        self.selected_files = []

        self.setup_ui()

    def setup_ui(self):
        """Thiết lập giao diện người dùng"""
        # Frame chính
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Tiêu đề
        title_label = ttk.Label(main_frame, text="Trích xuất CAS Number",
                                font=("Arial", 16, "bold"))
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 20))

        # Nút chọn file
        select_btn = ttk.Button(main_frame, text="Chọn file SDS",
                                command=self.select_files)
        select_btn.grid(row=1, column=0, sticky=tk.W, pady=5)

        # Label hiển thị số file đã chọn
        self.file_count_label = ttk.Label(
            main_frame, text="Chưa chọn file nào")
        self.file_count_label.grid(row=1, column=1, sticky=tk.W, padx=5)

        # Listbox hiển thị file
        self.file_listbox = tk.Listbox(main_frame, height=10)
        self.file_listbox.grid(
            row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)

        # Scrollbar cho listbox
        scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL,
                                  command=self.file_listbox.yview)
        scrollbar.grid(row=2, column=2, sticky=(tk.N, tk.S))
        self.file_listbox.configure(yscrollcommand=scrollbar.set)

        # Nút xóa file
        remove_btn = ttk.Button(main_frame, text="Xóa file đã chọn",
                                command=self.remove_selected_file)
        remove_btn.grid(row=3, column=0, sticky=tk.W, pady=5)

        # Progress bar - tăng tốc độ gấp 4 lần (interval = 12ms)
        self.progress = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress.grid(row=4, column=0, columnspan=2,
                           sticky=(tk.W, tk.E), pady=10)

        # Nút xử lý
        process_btn = ttk.Button(main_frame, text="Bắt đầu xử lý",
                                 command=self.process_files)
        process_btn.grid(row=5, column=0, columnspan=2, pady=10)

        # Cấu hình grid weights
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(2, weight=1)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

    def select_files(self):
        """Chọn file PDF"""
        files = filedialog.askopenfilenames(
            title="Chọn file SDS",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )

        if files:
            self.selected_files.extend(files)
            self.update_file_list()

    def remove_selected_file(self):
        """Xóa file đã chọn khỏi danh sách"""
        selected_indices = self.file_listbox.curselection()
        for index in selected_indices[::-1]:
            if index < len(self.selected_files):
                self.selected_files.pop(index)
        self.update_file_list()

    def update_file_list(self):
        """Cập nhật danh sách file"""
        self.file_listbox.delete(0, tk.END)
        for file_path in self.selected_files:
            self.file_listbox.insert(tk.END, os.path.basename(file_path))

        self.file_count_label.config(
            text=f"Đã chọn {len(self.selected_files)} file"
        )

    def process_files(self):
        """Xử lý tất cả file đã chọn"""
        if not self.selected_files:
            messagebox.showwarning(
                "Cảnh báo", "Vui lòng chọn ít nhất một file SDS!")
            return

        # Hiển thị progress bar với tốc độ nhanh gấp 4 lần (12ms thay vì 50ms)
        self.progress.start(12)

        # Chạy xử lý trong thread riêng để không block UI
        import threading
        thread = threading.Thread(target=self._process_files_thread)
        thread.daemon = True
        thread.start()

    def _process_files_thread(self):
        """Thread xử lý file"""
        try:
            # Xử lý file
            results = self.reader.process_multiple_files(self.selected_files)

            # Xuất ra Excel theo đường dẫn trong config
            success = self.reader.export_to_excel(results)

            if success:
                total_valid = sum(result['count_valid'] for result in results)
                total_invalid = sum(result['count_invalid']
                                    for result in results)
                total_cas = total_valid + total_invalid

                # Dùng after để hiển thị messagebox trong UI thread
                self.root.after(0, lambda: messagebox.showinfo(
                    "Thành công",
                    f"Đã xử lý xong!\n\n"
                    f"Tổng số file: {len(results)}\n"
                    f"Tổng số CAS Number tìm thấy: {total_cas}\n"
                    f"- CAS hợp lệ: {total_valid}\n"
                    f"- CAS không hợp lệ: {total_invalid}\n"
                    f"File đã được lưu tại: {config.EXCEL_SOURCE}"
                ))

        except Exception as e:
            # Dùng after để hiển thị messagebox trong UI thread
            self.root.after(0, lambda: messagebox.showerror(
                "Lỗi", f"Đã xảy ra lỗi: {str(e)}"))

        finally:
            # Dừng progress bar trong UI thread
            self.root.after(0, self.progress.stop)


def main():
    """Hàm chính"""
    root = tk.Tk()
    app = SDSApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()

